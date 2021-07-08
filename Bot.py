#imports libs
from tkinter import filedialog
from tkinter.constants import BOTTOM, LEFT, W
from selenium import webdriver
from webdriver_manager import driver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from tkinter import *
from time import sleep
import keyboard
import openpyxl as opyxl
import os

#Define Variable
css_selector = "#main > footer > div._3SvgF._1mHgA.copyable-area > div.DuUXI > div > div._1awRl.copyable-text.selectable-text"
mob_num=[]
img_add="test"
xlsx_add=""
msg="Hellow World!!!"
range_from=1
range_to=5
web_driver=""
wb=""
sh1=""

class Gui_interface:
    def __init__(self,master):
        # Title
        title = Label(master, text = "Whatappp Bot") 
        title.config(font =("Courier", 18)) 
        title.grid(row=0,columnspan=3)
        # Import_Xlsx Button
        openxlsx = Button(master, text = "Import Excel File",command = functions.add_Xlsx) 
        openxlsx.grid(row=1)
        t_1= Label(master,text="#open onlt Xlsx files")
        t_1.grid(row=1,column=1,columnspan=2)
        # Insert Message Box
        t_2= Label(master,text="Message:")
        t_2.grid(row=2)
        enter_message=Entry(master)
        enter_message.grid(row=2,column=1,rowspan=5,columnspan=2)
        # Image Checklist
        img_check = Checkbutton(master,text="Check If Image")
        img_check.grid(row=7)
        # Import Image
        open_img = Button(master, text = "Open Image",command = functions.add_img) 
        open_img.grid(row=7,column=1)
        # Img path
        t_3 = Label(master, text="Image Path-"+ img_add)
        t_3.grid(row=7,column=2)
        # Range
        t_4 = Label(master,text="Sr.no.")
        t_4.grid(row=8)
        enter_from=Entry(master)
        enter_from.grid(row=8,column=1)
        enter_to=Entry(master)
        enter_to.grid(row=8,column=2)
        #range_from_to = int(enter_to) - int(enter_from)
        #print(type(enter_to))
        #Start Button
        start_whatsapp = Button(master, text = "Start Whatsapp",command = functions.start_whatsapp) 
        start_whatsapp.grid(row=9)
        #Send Button
        send = Button(master, text = "Start Sent",command = functions.send_text) 
        send.grid(row=9,column=1)
        #Exit
        exit_end = Button(master, text = "Exit", 
                command = master.destroy) 
        exit_end.grid(row=9,column=2)


class functions:
    
    def add_Xlsx():  # Import Excel
        xlsxfile =filedialog.askopenfile(initialdir="/", title="Select Xlsx File", filetypes =(("Excel","*.xlsx"),("All Files","*.*")) )
        print("Xlsx-"+xlsxfile.name) 
        xlsx_add = xlsxfile.name
        xlsx_add = xlsx_add.replace( "/" , "\\")
        wb = opyxl.load_workbook(xlsx_add)
        sh1 = wb['ContactList']
        print(sh1.cell(3,1).value)
        print("list added")

    def add_img(): # Import Img
        imgfile =filedialog.askopenfile(initialdir="/", title="Select Img", filetypes =(("png","*.png"),("jpg",".jpg"),("All Files","*.*")) )
        img_add = imgfile
        img_add = img_add.replace( "/" , "\\")
        #print("img_add-"+str(img_add))
        print(imgfile) 
        print("Image added")

    def start_whatsapp(): # Initiate WhatsappWeb 
        print("Project Initiate")
        web_driver = webdriver.Chrome(ChromeDriverManager().install())
        web_driver.get("https://web.whatsapp.com")  # first call without delay in order to scan qr code
        keyboard.wait('`')
        print("QR Scaned")
        sleep(1)
        #web_driver.get("https://www.google.com")

    def send_text(): # Start Sending Text
        cache_count=1
        for i in range(range_from,range_to):
            url = "https://web.whatsapp.com/send?phone=91" +sh1.cell(i+2,2).value+ "&text=" +msg #sh1.cell(i+2,1).value
            web_driver.get(url)
            sleep(1)
            functions.send_enter(i)
            cache_count=cache_count+1
            if cache_count == 10:
                functions().clr_cache()
                cache_count=1
        print("Text Sending Started")

    def send_enter(i): # Send Enter Button
        for j in range(5): 
            try:
                print("Try -", j+1)
                web_driver.find_element_by_css_selector(css_selector).send_keys(Keys.RETURN)
                sh1.cell(i+2,3).value= "Sent"
                print("Sent")
                web_driver.execute_script("window.onbeforeunload = function() {};")
                sleep(8)
                break
            except:
                sleep(5)
                
            if j == 4:
                sh1.cell(i+2,3).value= "Failed"

    def send_img(): # Start Sending Text

        print("Img_Send")

    def clr_cache(): #Clear Cache & Cookies
        driver.delete_all_cookies() #CookiesClean
        #CacheClean

    def status_reset():
        for x in range(range_from,range_to):
            sh1.cell(x+2,3).value= "Pending"
    def savexlsx():
        wb.save(xlsx_add)

#Define Function
def initiate_gui():
    root=Tk()
    #root.geometry("750x800") 
    b=Gui_interface(root)
    #f_import = Frame(master)
    #f_info = Frame(master)
    #f_range = Frame(master)
    #f_buttons = Frame(master)
    #f_import.grid()

    root.mainloop()

initiate_gui()
functions().savexlsx()
   



